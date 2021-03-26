# -*-coding:utf8-*-
import logging

import openpyxl
import psycopg2
import datetime
from email.mime.multipart import MIMEMultipart
import smtplib, os
from email.mime.text import MIMEText
# BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# os.sys.path.append(BASE_DIR)



class check_purchase_order_cancel_config(object):
    """
    应取消po配置
    """
    def __init__(self):
        self.DATABASES_HOST = '192.168.1.90'
        # self.DATABASES_HOST = '47.106.200.13'
        self.DATABASES_NAME = "starmerx"
        self.DATABASES_USER = "wenhuang"
        self.DATABASES_PASSWORD = "Wen2012"
        self.DATABASES_PORT = '5432'

        self.EMAIL_USER = 'yf_system@starmerx.com'
        self.EMAIL_PASSWORD = 'Tianhu2017'

        #self.TO_EMAIL_LIST = ['guohaiqing@starmerx.com']
        self.TO_EMAIL_LIST = ['pengchenchen@starmerx.com', 'chaochao@starmerx.com', 'liangjianwei@starmerx.com', 'jinjiu@starmerx.com',
          'guohuan@starmerx.com', 'yucun@starmerx.com', 'zhengjie@starmerx.com', 'prod2@starmerx.com',
          'liuyanfang@starmerx.com', 'mingzhu@starmerx.com', 'fengaiping@starmerx.com', 'wuhui@starmerx.com',
          'lianchusheng@starmerx.com', 'xianfu@starmerx.com','mingliang@starmerx.com']

"""
1.重构询价单脚本应取消po
2.author:秦雨村
"""

# purchase_cancel_log = logging.getLogger('purchase_cancel')
# purchase_cancel_log.setLevel(level=logging.INFO)
# handler = logging.FileHandler("%s/qinyucun/log/purchase_cancel.log" %BASE_DIR)
# handler.setLevel(logging.INFO)
# formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
# handler.setFormatter(formatter)
# purchase_cancel_log.addHandler(handler)

class check_purchase_order_cancel(check_purchase_order_cancel_config):

    def export_info(self):
        data = datetime.datetime.today().strftime('%Y%m%d%H%M%S')
        file_path = 'purchase_cancel_%s.xlsx' % (data)
        excel = openpyxl.Workbook()

        ws = excel.create_sheet(title='purchase_cancel')
        self.write_sheet(ws)

        excel.save(file_path)
        self.send_email(file_path)


    def write_sheet(self,workbook):
        conn = psycopg2.connect(host=self.DATABASES_HOST, database=self.DATABASES_NAME, user=self.DATABASES_USER,
                                password=self.DATABASES_PASSWORD, port=self.DATABASES_PORT)
        cur = conn.cursor()

        should_be_cancelled_sql = """select a.product_id,a.sku,(a.stock_qty-a.coming_out_qty+a.route_qty) as future_qty,b.product_min_qty,b.product_max_qty
            from (
            select  a.id product_id,a.default_code sku
            ,sum(case when type  ='shelf' then stock_qty else 0 end )stock_qty
            ,sum(case when type  ='virtual' then stock_qty else 0 end )route_qty
            ,sum(case when type  ='delivery' then stock_qty else 0 end )coming_out_qty
             from product_product  a
             left join starmerx_inventory b on a.id=b.product_id
             where a.id != 2475 and a.product_tmpl_id in (select id from product_template where purchase_ok='t')  and b.location_id  in (193,182)
             group by a.id
             ) a left join stock_warehouse_orderpoint b on a.product_id=b.product_id
             where a.route_qty>0 and a.stock_qty-a.coming_out_qty+a.route_qty>b.product_max_qty;"""
        cur.execute(should_be_cancelled_sql)
        should_be_cancelled = cur.fetchall()



        workbook.cell(1, 1, 'SKU')
        workbook.cell(1, 2, '预测库存')
        workbook.cell(1, 3, '最小安全库存')
        workbook.cell(1, 4, '最大安全库存')
        workbook.cell(1, 5, 'PO')
        workbook.cell(1, 6, '采购数量')
        workbook.cell(1, 7, '应采购数量')
        workbook.cell(1, 8, '应取消数量')
        workbook.cell(1, 9, '采购员')
        workbook.cell(1, 10, '仓库状态')
        workbook.cell(1, 11, '付款状态')
        workbook.cell(1, 12, '跟踪单号')
        workbook.cell(1, 13, '该po里对应的sku数���')
        workbook.cell(1, 14, 'SKU单价')
        workbook.cell(1, 15, 'PO单合计金额')
        workbook.cell(1, 16, 'PO的供应商')
        workbook.cell(1, 17, 'SKU默认仓')


        i = 2
        count = 0
        for item in should_be_cancelled:
            product_id = item[0]
            sku = item[1]
            future_qty = item[2]
            product_min_qty = item[3]
            product_max_qty = item[4]

            sql1 = '''select a.name,a.stock_state,b.product_qty,b.should_purchase_qty_real,c.name,b.is_n3,a.location_id,a.id,b.move_dest_id,track_number,b.price_unit,a.amount_total,d.name,e.name from purchase_order a
                        left join purchase_order_line b on a.id=b.order_id
                        left join (select a.id,b.name from res_users a,res_partner b where a.partner_id=b.id) c on a.purchaser_id=c.id
                        left join res_partner d on a.partner_id=d.id
                        left join (select a.id,b.name from product_product a,stock_location b where a.mr_location_id=b.id) e on b.product_id=e.id
                        where not a.origin ~ 'S6' and not a.origin ~ 'S7' and (a.state in('draft') or a.stock_state in('assigned','receiving')) and b.product_id=%s
                        and b.product_qty>0 and b.product_qty<=b.should_purchase_qty_real and b.person_stock_id is null order by a.id desc limit 1''' % (
                product_id)

            cur.execute(sql1)
            result1 = cur.fetchone()

            if result1:
                po = result1[0]
                stock_state = result1[1].replace('none', '无入库单').replace('assigned', '准备收货').replace('receiving',
                                                                                                     '正在收货')
                product_qty = int(result1[2])
                should_purchase_qty_real = result1[3]
                name = result1[4]
                cancel_qty = future_qty - product_max_qty
                is_n3 = result1[5]
                location_id = result1[6]
                po_id = result1[7]
                move_dest_id = result1[8]
                track_number = str(result1[9])
                price = result1[10]
                price_total = result1[11]
                partner = result1[12]
                ware = result1[13]
                if is_n3 == '1' and cancel_qty <= 3:
                    continue

                sql3 = 'select count(*) from purchase_order_line where order_id=%s and product_id !=2475 group by product_id' % (
                    po_id)
                cur.execute(sql3)
                result3 = cur.fetchall()
                sku_count = len(result3)

                sql4 = """select string_agg(a.state,'+') from (select state from account_invoice where origin like '%%%s%%' and type='in_invoice' group by state) a""" % (
                    po)
                cur.execute(sql4)
                result4 = cur.fetchone()
                invoice_state = ''

                if result4 and result4[0]:
                    invoice_state = result4[0]
                    invoice_state = invoice_state.replace('draft', u'草稿')
                    invoice_state = invoice_state.replace('validate', u'已生效')
                    invoice_state = invoice_state.replace('open', u'已审核')
                    invoice_state = invoice_state.replace('paid', u'已付款').replace('partPaid', u'部分支付').replace('cancel',
                                                                                                               u'已取消')




                workbook.cell(row=i, column=1, value=sku)
                workbook.cell(row=i, column=2, value=future_qty)
                workbook.cell(row=i, column=3, value=product_min_qty)
                workbook.cell(row=i, column=4, value=product_max_qty)
                workbook.cell(row=i, column=5, value=po)
                workbook.cell(row=i, column=6, value=product_qty)
                workbook.cell(row=i, column=7, value=should_purchase_qty_real)
                workbook.cell(row=i, column=8, value=cancel_qty)
                workbook.cell(row=i, column=9, value=name)
                workbook.cell(row=i, column=10, value=stock_state)
                workbook.cell(row=i, column=11, value=invoice_state)
                workbook.cell(row=i, column=12, value=track_number)
                workbook.cell(row=i, column=13, value=sku_count)
                workbook.cell(row=i, column=14, value=price)
                workbook.cell(row=i, column=15, value=price_total)
                workbook.cell(row=i, column=16, value=partner)
                workbook.cell(row=i, column=17, value=ware)
                i += 1
                if stock_state == '无入库单':
                    count += 1
                    sql2 = "select id,stock_qty from starmerx_inventory where location_id=%s and type='virtual' and product_id=%s" % (
                        location_id, product_id)
                    cur.execute(sql2)
                    result2 = cur.fetchone()
                    inventory_id = result2[0]
                    stock_qty = int(result2[1])

                    if product_qty - cancel_qty > 0:
                        cur.execute(
                            "update purchase_order_line set write_date=(now() at time zone 'UTC'),product_qty=%s where order_id=%s and product_id=%s and product_qty>0",
                            (product_qty - cancel_qty, po_id, product_id))
                        cur.execute(
                            "update starmerx_inventory set write_date=(now() at time zone 'UTC'),stock_qty=%s where id=%s",
                            (stock_qty - cancel_qty, inventory_id))
                        cur.execute(
                            "update purchase_order a set amount_total=total,amount_untaxed=total from (select order_id,sum(price_unit*product_qty) total from purchase_order_line group by order_id) bwhere a.id=b.order_id and a.id=%s",
                            (po_id,))
                    else:
                        cur.execute(
                            "update stock_move set write_date=(now() at time zone 'UTC'),state='cancel' where id=%s",
                            (move_dest_id,))
                        cur.execute(
                            "update procurement_order set write_date=(now() at time zone 'UTC'),state='cancel' where move_id=%s",
                            (move_dest_id,))
                        cur.execute(
                            "update starmerx_inventory set write_date=(now() at time zone 'UTC'),stock_qty=%s where id=%s",
                            (stock_qty - product_qty, inventory_id))
                        if sku_count == 1:
                            cur.execute(
                                "update purchase_order set write_date=(now() at time zone 'UTC'),reason='应取消PO',state='cancel' where id=%s",
                                (po_id,))
                        else:
                            cur.execute("delete from purchase_order_line where order_id=%s and product_id=%s",
                                        (po_id, product_id))
                            cur.execute(
                                "update purchase_order a set amount_total=total,amount_untaxed=total from (select order_id,sum(price_unit*product_qty) total from purchase_order_line group by order_id) b where a.id=b.order_id and a.id=%s",
                                (po_id,))
                    conn.commit()
        cur.close()
        conn.close()



    def send_email(self,path):
        msg = MIMEMultipart()

        # 添加附件
        basename = os.path.basename(path)
        att1 = MIMEText(open(path, 'rb').read(), 'base64', 'gb2312')
        att1["Content-Type"] = 'application/octet-stream'
        att1["Content-Disposition"] = 'attachment; filename=%s' % basename
        msg.attach(att1)
        # 加邮件头
        msg['to'] = ';'.join(self.TO_EMAIL_LIST)
        msg['from'] = self.EMAIL_USER
        msg['subject'] = '不可采购或不可销售应取消的PO'
        # 发送邮件
        server = smtplib.SMTP()
        server.connect('smtp.exmail.qq.com')
        server.login(self.EMAIL_USER, self.EMAIL_PASSWORD)
        server.sendmail(msg['from'], self.TO_EMAIL_LIST, msg.as_string())
        server.quit()
        # 退出后删除xls文件
        os.remove(path)


if __name__ == '__main__':
    print('=========应取消po脚本执行开始===========')
    check_obj = check_purchase_order_cancel()
    check_obj.export_info()
    print('=========应取消po脚本执行结束===========')